import streamlit as st
import pandas as pd
import json
import os
from io import BytesIO
from openpyxl import load_workbook

# ---------------- Utils ----------------
def load_dataframe(file, sheet_name=None):
    filename = file.name.lower() if hasattr(file, "name") else file.lower()

    if filename.endswith(".csv"):
        df = pd.read_csv(file)
    elif filename.endswith(".xlsx"):
        if sheet_name is None:
            raise ValueError("sheet_name is required for Excel files")
        
        # Use openpyxl to read the actual data and find max column
        if hasattr(file, 'read'):
            file.seek(0)
            wb = load_workbook(file)
        else:
            wb = load_workbook(file)
        
        ws = wb[sheet_name]
        
        # Read the data manually to get proper headers from row 1 and row 2
        headers = []
        data = []
        
        # Read header rows (1 and 2)
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=2, values_only=True), 1):
            # Only keep columns up to the max used column
            row = row[:ws.max_column]
            if row_idx == 1:
                row1 = row
            elif row_idx == 2:
                row2 = row
        
        # Combine headers from row 1 and row 2
        for i, (h1, h2) in enumerate(zip(row1, row2)):
            h1_str = str(h1).strip() if h1 else ""
            h2_str = str(h2).strip() if h2 else ""
            
            # Combine headers
            if h1_str and h2_str and h1_str != h2_str:
                combined = f"{h1_str} - {h2_str}"
            elif h1_str:
                combined = h1_str
            elif h2_str:
                combined = h2_str
            else:
                combined = f"Column_{i}"
            
            headers.append(combined)
        
        # Read data rows (skip header rows)
        for row in ws.iter_rows(min_row=3, values_only=True):
            row = row[:ws.max_column]
            data.append(row)
        
        wb.close()
        
        # Create dataframe
        df = pd.DataFrame(data, columns=headers)
        
        # Filter out invalid columns
        valid_columns = []
        for col in df.columns:
            col_str = str(col).strip()
            if (col_str and 
                'unnamed' not in col_str.lower() and 
                col_str.lower() != 'nan' and
                col_str != '' and
                not col_str.startswith('Column_')):
                valid_columns.append(col)
        
        df = df[valid_columns]
        
        # Make column names unique by adding counter if duplicates exist
        seen = {}
        unique_cols = []
        for col in df.columns:
            if col in seen:
                seen[col] += 1
                unique_cols.append(f"{col}_{seen[col]}")
            else:
                seen[col] = 0
                unique_cols.append(col)
        df.columns = unique_cols
    else:
        raise ValueError("Unsupported file type")
    
    return df

# ---------------- App Config ----------------
st.set_page_config(page_title="Smart Excel Merger", layout="wide")
st.title("📊 Smart Excel Merger with Column Mapping")

BASE_DIR = os.getcwd()
DATA_DIR = os.path.join(BASE_DIR, "data")
MAPPING_DIR = os.path.join(BASE_DIR, "mappings")

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(MAPPING_DIR, exist_ok=True)

TEMPLATE_PATH = os.path.join(DATA_DIR, "template.xlsx")

# ---------------- Session State ----------------
st.session_state.setdefault("column_mapping", {})
st.session_state.setdefault("merge_keys", [])
st.session_state.setdefault("mapping_confirmed", False)

# ---------------- Template ----------------
st.subheader("📌 Template File")

template_ready = False

if os.path.exists(TEMPLATE_PATH):
    st.success("✅ Template file is set")
    template_ready = True
    if st.button("🔄 Replace Template"):
        os.remove(TEMPLATE_PATH)
        st.session_state.clear()
        st.rerun()
else:
    st.warning("No template file found. Please upload one to get started.")
    uploaded_template = st.file_uploader("Upload Template Excel", type=["xlsx"])
    if uploaded_template:
        with open(TEMPLATE_PATH, "wb") as f:
            f.write(uploaded_template.getbuffer())
        st.success("Template saved!")
        st.rerun()

if not template_ready:
    st.stop()

# ---------------- Upload File B ----------------
file_b = st.file_uploader("Upload Excel or CSV File B", type=["xlsx", "csv"])
if not file_b:
    st.stop()

# ---------------- Read Template ----------------
excel_a = pd.ExcelFile(TEMPLATE_PATH)
sheet_a = st.selectbox("Select sheet from Template", excel_a.sheet_names)
df_a = load_dataframe(TEMPLATE_PATH, sheet_name=sheet_a)

# ---------------- Read File B ----------------
if file_b.name.lower().endswith(".xlsx"):
    excel_b = pd.ExcelFile(file_b)
    sheet_b = st.selectbox("Select sheet from File B", excel_b.sheet_names)
    df_b = load_dataframe(file_b, sheet_name=sheet_b)
else:
    st.info("CSV detected — no sheet selection needed")
    df_b = load_dataframe(file_b)

# ---------------- Row Count ----------------
st.subheader("📈 Row Count (Before Merge)")
c1, c2 = st.columns(2)
c1.metric("Template Rows", len(df_a))
c2.metric("File B Rows", len(df_b))

# ---------------- Load Mapping ----------------
st.subheader("💾 Load Column Mapping (Optional)")
mapping_files = [f for f in os.listdir(MAPPING_DIR) if f.endswith(".json")]
selected_mapping = st.selectbox("Select saved mapping", ["None"] + mapping_files)

if selected_mapping != "None":
    with open(os.path.join(MAPPING_DIR, selected_mapping)) as f:
        saved = json.load(f)
        st.session_state.column_mapping = saved["column_mapping"]
        st.session_state.merge_keys = saved["merge_keys"]
        st.session_state.mapping_confirmed = True
    st.success(f"Loaded mapping: {selected_mapping}")

# ---------------- Column Mapping (FORM) ----------------
st.subheader("🧩 Column Mapping (Template → File B)")

# Show helpful info about the columns
col1, col2 = st.columns(2)
with col1:
    st.info(f"📋 Template has {len(df_a.columns)} columns: {', '.join(df_a.columns.tolist())}")
with col2:
    st.info(f"📋 File B has {len(df_b.columns)} columns: {', '.join(df_b.columns.tolist())}")

st.markdown("---")
st.markdown("**Map each Template column to a File B column (or Ignore it):**")

with st.form("column_mapping_form"):
    temp_mapping = {}
    temp_keys = []

    # Create a container for the mappings
    with st.container():
        # Display column mappings with better clarity
        for col_idx, col in enumerate(df_a.columns):
            st.markdown(f"**Template Column: `{col}`**")
            col_map, col_key, col_preview = st.columns([2, 1, 1])

            default = st.session_state.column_mapping.get(col, "Ignore")

            with col_map:
                selected = st.selectbox(
                    f"Maps to",
                    ["Ignore"] + list(df_b.columns),
                    key=f"map_{col_idx}_{col}",  # Add index to ensure uniqueness
                    index=(["Ignore"] + list(df_b.columns)).index(default)
                    if default in ["Ignore"] + list(df_b.columns) else 0,
                    label_visibility="collapsed"
                )

            with col_key:
                is_key = st.checkbox(
                    "Merge Key?",
                    key=f"merge_key_{col_idx}_{col}",  # Add index to ensure uniqueness
                    value=col in st.session_state.merge_keys
                )

            with col_preview:
                if selected != "Ignore":
                    sample_a = str(df_a[col].dropna().iloc[0] if len(df_a[col].dropna()) > 0 else "")[:20]
                    sample_b = str(df_b[selected].dropna().iloc[0] if len(df_b[selected].dropna()) > 0 else "")[:20]
                    st.caption(f"Sample: '{sample_a}' → '{sample_b}'")

            if selected != "Ignore":
                temp_mapping[col] = selected
                if is_key:
                    temp_keys.append(col)
            
            st.markdown("")  # Add spacing between rows

    submitted = st.form_submit_button("✅ Confirm Mapping")

if submitted:
    if not temp_mapping:
        st.error("Map at least one column.")
        st.stop()

    st.session_state.column_mapping = temp_mapping
    st.session_state.merge_keys = temp_keys
    st.session_state.mapping_confirmed = True
    st.success("Mapping confirmed!")

# ---------------- Save Mapping ----------------
if st.session_state.mapping_confirmed:
    st.subheader("💾 Save Mapping")
    mapping_name = st.text_input("Mapping name")

    if st.button("Save Mapping"):
        if not mapping_name:
            st.error("Please enter a mapping name.")
        else:
            with open(os.path.join(MAPPING_DIR, f"{mapping_name}.json"), "w") as f:
                json.dump(
                    {
                        "column_mapping": st.session_state.column_mapping,
                        "merge_keys": st.session_state.merge_keys,
                    },
                    f,
                    indent=2,
                )
            st.success("Mapping saved!")

# ---------------- Merge ----------------
if st.session_state.mapping_confirmed:
    st.subheader("🔀 Fill Template with Data")

    if st.button("🚀 Fill Template"):
        # Create a copy of the template to fill
        filled_df = df_a.copy()
        
        # Fill the first column with serial numbers
        if len(filled_df.columns) > 0:
            first_column = filled_df.columns[0]
            filled_df[first_column] = range(1, len(filled_df) + 1)
        
        # For each mapped column, copy data from File B to template
        for template_col, file_b_col in st.session_state.column_mapping.items():
            if file_b_col in df_b.columns:
                # Get the data from File B (handle different row counts)
                if len(df_b) >= len(filled_df):
                    # File B has more or equal rows - take as many as we need
                    filled_df[template_col] = df_b[file_b_col].iloc[:len(filled_df)].values
                else:
                    # File B has fewer rows - fill what we can, rest remains empty
                    filled_df[template_col] = pd.NA
                    filled_df.loc[:len(df_b)-1, template_col] = df_b[file_b_col].values

        st.subheader("📊 Template Filled with Data")
        m1, m2 = st.columns(2)
        m1.metric("Template Rows", len(df_a))
        m2.metric("Mapped Columns Updated", len(st.session_state.column_mapping))

        st.dataframe(filled_df.head(50))

        output = BytesIO()
        filled_df.to_excel(output, index=False)
        output.seek(0)

        st.download_button(
            "⬇️ Download Filled Template",
            output,
            "filled_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
